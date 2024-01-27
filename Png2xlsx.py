## using openpyxl to change background color of each cell in a worksheet so that an image is deplicated from a PNG photo

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image
import os

# change color from rgb value to hex value
def rgb_to_hex(rgb):
    return f"{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"

filename = 'job.xlsx'  # Replace with your desired file name

# Check if the file exists, and create it if it doesn't
if not os.path.exists(filename):    
    wb = Workbook()
    wb.save(filename)
    print('New Excel file opened successfully.')
else:
    # Open the existing Excel file
    wb = load_workbook(filename)
    print('Existing Excel file opened successfully.')

# Get the sheet titles
sheet_titles = wb.sheetnames

# Iterate through each sheet
# remove all other worksheet except the one just created
print('Removing other worksheets ....')
for sheet_title in sheet_titles:
    # if sheet_title != '1':
        # Delete the sheet
    wb.remove(wb[sheet_title])

# Create a list of worksheet with titles 1,2,3,...
worksheet_titles = [f'{i}' for i in range(1, 2)]   #one worksheet for this application

# Open each worksheet and do something with it
for title in worksheet_titles:
    print('Opening worksheet '+title)
    wb.create_sheet(title)
    worksheet = wb[title]
    worksheet.sheet_view.showGridLines = False

    # Ask user to input png file name
    image_name = input("Please enter your file name (without the file type in the end): ")
    file_type = ".png"
    image_name_with_file_type = image_name + file_type
    
    # Open the PNG file
    image = Image.open(image_name_with_file_type)
    
    # Get the size of the image
    width, height = image.size
    
    # Define the range of cells to format
    start_row = 1
    end_row = height
    start_col = 1
    end_col = width
    
    # [To do] End_row and end_col replaced from the PNG image...
    
    print('Changing cell background color ....')
    # Set the column widths
    for col in range(start_col, end_col+1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 0.268

    # Set the row heights
    for row in range(start_row, end_row+1):
        worksheet.row_dimensions[row].height = 2
        # Fill pixel into the cell using pixel from png fle (one pixel to one cell copy) 
        for col in range(start_col, end_col+1):
            cell = worksheet.cell(row=row, column=col)
            # Get the pixel color at coordinates (x,y)
            pixel_color = image.getpixel((col-1, row-1))   #(255, 255, 255)
            # Create an RGBColor object from the separate variables
            color = rgb_to_hex(pixel_color)   # FFFFFF
            # Set the fill color of the cell to red
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")



# Save the workbook
print('Saving Excel file ....')
wb.save(filename)

# Close the workbook
print('Closing Excel file.')
wb.close()

# Open the Excel file with Office Excel application
print('Opening Excel file using Office Excel.')
os.startfile(filename)